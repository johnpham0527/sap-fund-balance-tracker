### Helper Functions
def getFundValue(cell, row):
	return cell(row=row, column=7).value

def getReleasedBudgetValue(cell, row):
	return cell(row=row, column=26).value

def getYTDActualValue(cell, row):
	return cell(row=row, column=29).value

def getOpenPOsValue(cell, row):
	return cell(row=row, column=34).value

def getOpenReqsValue(cell, row):
	return cell(row=row, column=37).value

def getBalanceValue(cell, row):
	return cell(row=row, column=42).value

def getFundNameRow(cell, row):
	currentRow = row
	currentFundCell = inputSheet.cell(row=currentRow, column=2) 
	while currentFundCell.value != "Fund Name:": #search for the row which contains "Fund Name:"
		#move to the previous row until "Fund Name:" is found
		currentRow -= 1 
		currentFundCell = inputSheet.cell(row=currentRow, column=2)
	return currentRow

def getEndDate(cell, row):
	return cell(row=row, column=39).value

def getFundName(cell, row):
	return cell(row=row, column=4).value


### Import
import openpyxl
import datetime
from openpyxl import Workbook

### Initiate input variables
inputWorkbook = openpyxl.load_workbook('AllFunds.xlsx')
inputSheet = inputWorkbook['AllFunds']
inputSheetMaxRows = inputSheet.max_row

### Initiate output variables
outputWorkbook = Workbook()
outputArray = []

### Iterate through the input spreadsheet
print("Reading through the AllFunds.xlsx spreadsheet...")
for i in range(1,inputSheetMaxRows-2,1):
	currentFundCell = inputSheet.cell(row=i, column=7)
	if currentFundCell.value is not None: #if Column 7 is not empty, then we have found a fund
		#retrieve values for the fund
		currentRow = i #the fund value, released budget, YTD actual, open POs, open Reqs, and balance are all found in the current row
		fund = getFundValue(inputSheet.cell, currentRow)
		releasedBudget = getReleasedBudgetValue(inputSheet.cell, currentRow)
		ytdActual = getYTDActualValue(inputSheet.cell, currentRow)
		openPOs = getOpenPOsValue(inputSheet.cell, currentRow)
		openReqs = getOpenReqsValue(inputSheet.cell, currentRow)
		balance = getBalanceValue(inputSheet.cell, currentRow)
		currentRow = getFundNameRow(inputSheet.cell, currentRow) #the fund name and end date are found in a different row that contains "Fund Name:"
		endDate = getEndDate(inputSheet.cell, currentRow)
		fundName = getFundName(inputSheet.cell, currentRow)
		rowArray = [fund, releasedBudget, ytdActual, openPOs, openReqs, balance, endDate, fundName]
		outputArray.append(rowArray)

### Populate header row of output worksheet
outputSheet = outputWorkbook.active
outputSheet['A1'] = "Fund"
outputSheet['B1'] = "Released Budget"
outputSheet['C1'] = "YTD Actual"
outputSheet['D1'] = "Open POs"
outputSheet['E1'] = "Open Reqs"
outputSheet['F1'] = "Balance"
outputSheet['G1'] = "End Date"
outputSheet['H1'] = "Fund Name"

### Populate output worksheet with outputArray values
outputRow = 2
print("Outputting the cleaned values into a new spreadsheet...")
for fundRow in outputArray:
	outputSheet.cell(row=outputRow, column=1, value = fundRow[0]) #Fund
	outputSheet.cell(row=outputRow, column=2, value = fundRow[1]) #Released Budget
	outputSheet.cell(row=outputRow, column=3, value = fundRow[2]) #YTD actual
	outputSheet.cell(row=outputRow, column=4, value = fundRow[3]) #Open POs
	outputSheet.cell(row=outputRow, column=5, value = fundRow[4]) #Open Reqs
	outputSheet.cell(row=outputRow, column=6, value = fundRow[5]) #Balance
	outputSheet.cell(row=outputRow, column=7, value = fundRow[6]) #End Date
	outputSheet.cell(row=outputRow, column=8, value = fundRow[7]) #Fund Name
	outputRow = outputRow + 1

### Save output workbook
outputWorkbook.save(filename = 'outputFunds.xlsx')
print("Done. Saved file as outputFunds.xlsx. Happy grants management!")